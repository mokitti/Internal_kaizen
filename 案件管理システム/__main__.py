import PySimpleGUI as sg
import configparser
from pathlib import Path
import openpyxl
import pyqrcode
import gui


def main():
    # 最初に表示するウィンドウを指定する
    window = gui.main_window()    
    
    try:
        # 設定ファイルを読み込む
        config = configparser.ConfigParser()
        config.read('config.ini', 'utf-8')
        # 初期設定が完了しているか確認
        caselist_path = config['setting']['caselist_path']
        # 新規登録と案件管理ボタンを有効化
        window['-Move_change_window-'].update(disabled=False)
        window['-Move_registration_window-'].update(disabled=False)
    except Exception as e:
        sg.popup('初期設定を行ってください。', e)

    
    # 各種イベント処理
    while True:
        event, values = window.read()
        
        if event == sg.WIN_CLOSED or event == 'Exit' or event == '終了':
            break

        # --------------------------------------------
        # メイン画面のイベント
        # --------------------------------------------
        # 初期設定画面へ移動
        if event == '初期設定':
            window.close()
            window = gui.initialize_window()

        # 新規登録画面に移動
        if event == '-Move_registration_window-':
            window.close()
            window = gui.registration_window()

        # 案件管理画面へ移動
        if event == '-Move_change_window-':
            window.close()
            window = gui.change_window()
            # 案件一覧.xlsxから案件管理番号を取得
            Case_id_list = get_Case_id_list(caselist_path)
            # 案件管理番号リストボックスに反映
            window['-Case_id_list-'].update(values=Case_id_list)

        # --------------------------------------------
        # 初期設定画面のイベント
        # --------------------------------------------
        # データ保存先が選択された場合、初期設定ボタンを有効化
        if event == '-Folder_path-':
            window['-Initialize-'].update(disabled=False)

        # アプリの初期設定（必要なフォルダ、ファイル類の作成と設定ファイル更新）
        if event == '-Initialize-':
            # データの保存先を取得
            folder_path = values['-Folder_path-']
            # アプリの初期設定
            initialize_app(folder_path)
            # 初期設定画面をクローズし、メイン画面へ切り替え
            window.close()
            window = gui.main_window()

            # 新規登録、案件管理ボタンを有効化
            window['-Move_change_window-'].update(disabled=False)
            window['-Move_registration_window-'].update(disabled=False)


        # --------------------------------------------
        # 新規登録画面のイベント
        # --------------------------------------------
        # 案件の新規登録
        if event == '-Register-':
            try:
                # ウィジェットの値を取得
                input_data = [values['-Case_id-'], values['-Case_name-'],
                    values['-Shelf_number-'], values['-Owner-']]
                # 案件管理番号がない場合はValueErrorを発生
                if input_data[0] == '':
                    raise ValueError('案件管理番号を入力して下さい。')
                # 案件管理番号に重複が無いかをチェック
                if is_unique_id(caselist_path, input_data[0]) is not True:
                    raise ValueError('案件管理番号が重複しています。別の番号を入力下さい。')

                # 案件一覧.xlsxに案件データを登録
                create_data(caselist_path, input_data, qrcode_path)
            except Exception as e:
                sg.popup(e)

        # --------------------------------------------
        # 案件管理画面のイベント
        # --------------------------------------------
        # 案件管理番号が選択されたとき、各ウィジェットに案件情報を反映
        if event == '-Case_id_list-':
            # 選択された案件管理番号を取得
            Case_id = values['-Case_id_list-'][0]
            # ウィジェットをアップデート
            wgt_update(caselist_path, Case_id, window)
        # 案件情報の更新
        if event == '-Update-':
            try:
                # ウィジェットの値をリスト型変数input_dataに格納
                Case_id = values['-Case_id_list-'][0]
                input_data = [Case_id, values['-Case_name-'],
                    values['-Shelf_number-'], values['-Owner-']]
                # 案件情報を更新
                update_data(caselist_path, input_data)
            except Exception as e:
                sg.popup(e)
        # 案件情報の削除
        if event == '-Delete-':
            try:
                # 削除対象の案件管理番号をdelete_Case_idに格納
                delete_Case_id = values['-Case_id_list-'][0]
                # 案件情報を削除
                delete_data(caselist_path, delete_Case_id, window)
            except Exception as e:
                sg.popup(e)

    window.close()

def initialize_app(path):
    # ディレクトリの作成
    folder_path = path
    # 案件フォルダの作成
    root_path = folder_path + '/案件管理'
    Path(root_path).mkdir()
    # 案件一覧.xlsxを作成
    caselist_path = root_path + '/案件一覧.xlsx'
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = '案件シート'
    # 列幅の設定
    column_widths = {'A': 16, 'B': 16, 'C': 16, 'D': 16}
    # 列の見出しの設定
    column_name = {
        'A': '案件管理番号', 'B': '品名', 'C': '棚番号', 'D': '管理者'}
    for count, column_index in enumerate(column_widths):
        count += 1
        sheet.column_dimensions[column_index].width = column_widths[column_index]
        sheet.cell(1, count).value = column_name[column_index]
    wb.save(caselist_path)
    # 設定ファイルconfig.iniを更新
    config = configparser.ConfigParser()
    config.read('config.ini', 'utf-8')
    # 設定ファイルにsettingセクションが無い場合のみセクション追加
    if config.has_section('setting') is not True:
        config.add_section('setting')
        config['setting']['root_path'] = root_path
        config['setting']['caselist_path'] = caselist_path
    configpath = Path('config.ini')
    with configpath.open(mode='w', encoding='utf-8') as f:
        config.write(f)

    return sg.popup('初期設定が完了しました。')


def is_unique_id(path, search_value):
    # 案件一覧.xlsxのデータを取得
    wb, ws, xlsxdata = get_xlsx_data(path)
    # 案件一覧.xlsxに重複する案件管理番号があるかチェック
    if search_row(xlsxdata, search_value) is None:
        # 重複なければ、True（重複なし）を返す
        return True
    else:
        return False


def get_xlsx_data(path):
    # 案件一覧.xlsxと案件シートを読み込み
    wb = openpyxl.load_workbook(path)
    ws = wb['案件シート']
    # 案件一覧.xlsxのデータを変数dataに格納
    last_address = 'D' + str(ws.max_row)
    data = ws['A2:{}'.format(last_address)]
    return wb, ws, data


def search_row(data, search_value):
    # whileで繰り返し検索
    row = 0
    while row < len(data):
        # 案件一覧.xlsxから一致する案件管理番号を検索
        if str(search_value) == str(data[row][0].value):
            # 一致する行番号は、row + 2
            matched_row = row + 2
            return matched_row
        else:
            row += 1
    else:
        # 一致しない場合は、Noneを返す
        return None

def create_data(path, data, qr_path):
    # 案件一覧.xlsxと案件シートを読み込み
    wb = openpyxl.load_workbook(path)
    ws = wb['案件シート']
    # 追記する行番号を設定
    last_row = ws.max_row + 1
    # セルにデータを書き込み
    for colm in range(len(data)):
        ws.cell(row=last_row, column=colm + 1, value=data[colm])
    # wbを上書き保存
    wb.save(path)
    return sg.popup_ok('書き込みが完了しました')


def get_Case_id_list(path):
    # 案件一覧.xlsxのデータを取得
    wb, ws, data = get_xlsx_data(path)
    # 案件管理番号を格納するリスト型変数Case_id_listを宣言
    Case_id_list = []
    # 案件一覧.xlsx内の案件管理番号をすべてCase_id_listへ格納
    for row in range(len(data)):
        Case_id_list.append(data[row][0].value)
    return Case_id_list


def wgt_update(path, search_value, window):
    # 案件一覧.xlsxのデータを取得
    wb, ws, xlsxdata = get_xlsx_data(path)
    # 案件一覧.xlsxから一致する案件管理番号を検索
    for row in range(len(xlsxdata)):
        # 案件管理番号が一致した場合の処理
        if str(search_value) == str(xlsxdata[row][0].value):
            # ウィジェットを一旦クリア
            window['-Case_name-'].update(value='')
            window['-Owner-'].update(value='')
            # ウィジェットを更新
            window['-Case_name-'].update(value=xlsxdata[row][1].value)
            window['-Owner-'].update(value=xlsxdata[row][3].value)
            break


def update_data(path, data):
    # 案件一覧.xlsxのデータを取得
    wb, ws, xlsxdata = get_xlsx_data(path)
    # 案件一覧.xlsxから一致する案件管理番号を検索し、行番号を取得
    matched_row = search_row(xlsxdata, data[0])
    # セルに値を上書き（更新）
    for colm in range(len(data)):
        ws.cell(row=matched_row, column=colm + 1, value=data[colm])
    # wbを上書き保存
    wb.save(path)

def delete_data(path, search_value, window):
    # 案件一覧.xlsxのデータを取得
    wb, ws, xlsxdata = get_xlsx_data(path)
    # 案件一覧.xlsxから一致する案件管理番号を検索し、行番号を取得
    matched_row = search_row(xlsxdata, search_value)
    # wsから削除対象の行を削除
    ws.delete_rows(matched_row)
    # wbを上書き保存
    wb.save(path)
    # リストボックスを更新
    Case_id_list = get_Case_id_list(path)
    window['-Case_id_list-'].update(values=Case_id_list)
    window['-Case_name-'].update(value='')
    window['-Owner-'].update(value='')

    return sg.popup('削除しました')






if __name__ == '__main__':
    main()