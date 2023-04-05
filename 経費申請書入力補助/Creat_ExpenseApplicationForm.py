import os
import openpyxl as xl
from datetime import datetime, timedelta

# ******* Function *************************************************************************************************************
def SerialToDateTime(serialVal:float) -> str:
    """日付シリアル値からyyyy/MM/dd HH:MM:ss 形式に変換
    Args:
        serialVal (float): シリアル値(Ex: 44434.3412172569)
    Returns:
        str: 時刻  yyyy/MM/dd HH:MM:ss 形式
    """      
    sDateTime = (dt(1899,12,30) + timedelta(serialVal)).strftime('%Y/%m/%d')

    return sDateTime

def get_first_date(dt):
    return dt.replace(day=1)

# ****** Main *******************************************************************************************************************

# 作業ディレクトリをPythonファイルがある場所に変更する
current_dir = os.path.dirname(__file__)
os.chdir(current_dir)
print(current_dir)

# 各Path設定
input_file = os.path.join(current_dir, "入力用.xlsx")
input_wb = xl.load_workbook(input_file, data_only=True)
input_ws = input_wb['Sheet1']
format_file = os.path.join(current_dir, f"経費精算書_{datetime.now().strftime('%Y年%m月分')}_名前.xlsx")
format_wb = xl.load_workbook(format_file)
format_ws = format_wb['精算書']
format_row = 14

# INPUTファイルの内容を取得し、フォーマットファイルに転記していく
for row in input_ws.iter_rows(min_row=3, min_col=3, max_row=input_ws.max_row, max_col=11):
    start_date = datetime.strptime(SerialToDateTime(row[0].value), '%Y/%m/%d')
    end_date = datetime.strptime(SerialToDateTime(row[1].value), '%Y/%m/%d')

    days = (end_date - start_date).days + 1 if start_date != end_date else 1

    for i in range(days):
        format_ws.cell(format_row, 2).value = start_date + timedelta(days=i)
        for j in range(2, 9):
            format_ws.cell(format_row, j+1).value = row[j].value
        format_row += 1

# C3セルの値に実行日日付を入力する
first_date = SerialToDateTime(input_ws["C3"].value)
format_ws["B11"].value = datetime.today()

# 所定のファイル名に変更し保存
output_file = os.path.join(current_dir, f"経費精算書_{first_date.strftime('%Y年%m月分')}_名前.xlsx")
format_wb.save(output_file)
